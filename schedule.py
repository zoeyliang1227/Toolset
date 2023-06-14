import openpyxl

# file_name = str(input('請輸入檔案名：'))
wb = openpyxl.load_workbook('專櫃值班表new-v2 - 複製.xlsx' , data_only=True)     # 開啟 Excel 檔案, 設定 data_only=True 只讀取計算後的數值
total = wb.create_sheet("六日加班(全)", 0)   # 插入工作表 0 在第一個位置
total.cell(row=1, column= 1).value = '值班人員'
total.cell(row=1, column= 2).value = '六日加班(全)'
dic = {}

def get_values():
    for sheet in wb.worksheets:
        rows = sheet.iter_rows(min_col=2, max_col=4, min_row=1, max_row=sheet.max_row)
        fir_count = 0
        sec_count = 0
        for row in rows:
            if row[0].value == '週六' or row[0].value == '週日':
                if row[1].value == '全':
                    dic.clear()
                    fir_count += 1
                if row[2].value == '全':
                    dic.clear()
                    sec_count += 1

        dic[sheet.cell(row=3, column= 3).value]=fir_count
        dic[sheet.cell(row=3, column= 4).value]=sec_count
        print(sheet, dic)
        for key, val in dic.items():
            total.append([key, val])
    wb.save('專櫃值班表new-v2 - 複製.xlsx')

get_values()